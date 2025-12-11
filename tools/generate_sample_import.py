import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()
s = wb.active
s.title = 'Matches'

# Header
headers = ['Season', 'Home Team', 'Away Team', 'Match Date', 'Matchday', 'Venue']
s.append(headers)

# Row 1: Match Date as string with date+time
s.append([1, 'ABAM', 'AMIGBO', '2025-02-12 16:00', 1, 'Main Field'])

# Row 2: Match Date as Excel datetime (datetime object)
s.append([1, 'ABAM', 'AMIGBO', datetime(2025, 2, 13, 18, 30), 2, 'Main Field'])

# Row 3: Invalid row (missing date)
s.append([1, 'ABAM', 'AMIGBO', '', 3, 'Main Field'])

# Optionally set column widths
for i, h in enumerate(headers, start=1):
    s.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

fname = 'sample_match_import.xlsx'
wb.save(fname)
print(f'Wrote sample file: {fname}')
