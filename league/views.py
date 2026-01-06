from rest_framework import status, viewsets
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, render
from .models import Team, Season, Match, News, Group, TeamGroup
from .serializers import TeamSerializer, SeasonSerializer, MatchSerializer, MatchCreateSerializer, NewsSerializer
from .permissions_groups import IsNewsUploaderOrReadOnly, IsResultsEditor
from .utils import compute_standings
from rest_framework.permissions import IsAuthenticated
import difflib
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
import openpyxl
import random
from django.utils import timezone

# List all teams for a given season
@api_view(['GET'])
def teams_for_season(request):
    season_id = request.query_params.get('season')
    if not season_id:
        return Response({'error': 'season parameter required'}, status=status.HTTP_400_BAD_REQUEST)
    teams = Team.objects.filter(seasongroup__season_id=season_id).distinct()
    # fallback: if no TeamGroup, show all teams for season
    if not teams.exists():
        teams = Team.objects.filter(season_id=season_id)
    data = [{'id': t.id, 'name': t.name} for t in teams]
    return Response(data)

# Add or remove a team from a group
@csrf_exempt
@api_view(['POST', 'DELETE'])
def group_team_modify(request):
    team_id = request.data.get('team_id')
    group_id = request.data.get('group_id')
    season_id = request.data.get('season_id')
    if not all([team_id, group_id, season_id]):
        return Response({'error': 'team_id, group_id, and season_id required'}, status=status.HTTP_400_BAD_REQUEST)
    team = get_object_or_404(Team, id=team_id)
    group = get_object_or_404(Group, id=group_id, season_id=season_id)
    if request.method == 'POST':
        TeamGroup.objects.update_or_create(team=team, group=group, season_id=season_id)
        return Response({'success': True, 'action': 'added', 'team': team.name, 'group': group.name})
    elif request.method == 'DELETE':
        TeamGroup.objects.filter(team=team, group=group, season_id=season_id).delete()
        return Response({'success': True, 'action': 'removed', 'team': team.name, 'group': group.name})
@api_view(['GET'])
def groups_for_season(request):
    season_id = request.query_params.get('season')
    if not season_id:
        return Response({'error': 'season parameter required'}, status=400)
    groups = Group.objects.filter(season_id=season_id)
    data = [{'id': g.id, 'name': g.name} for g in groups]
    return Response(data)
from django.shortcuts import render


class ManualTeamGroupView(APIView):
    def post(self, request):
        team_name = request.data.get('team_name')
        category = request.data.get('category')
        season_id = request.data.get('season_id')
        group_name = request.data.get('group_name')

        if not all([team_name, category, season_id, group_name]):
            return Response({'error': 'All fields are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            season = Season.objects.get(id=season_id, category=category)
        except Season.DoesNotExist:
            return Response({'error': 'Season not found.'}, status=status.HTTP_404_NOT_FOUND)

        group, _ = Group.objects.get_or_create(name=group_name, season=season)
        team, _ = Team.objects.get_or_create(name=team_name, defaults={'category': category})
        TeamGroup.objects.update_or_create(team=team, group=group)
        return Response({'success': True, 'team': team_name, 'group': group_name})

class ExcelImportView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, format=None):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file uploaded'}, status=400)
        try:
            wb = openpyxl.load_workbook(file_obj)
            sheet = wb.active
        except Exception as e:
            return Response({'error': f'Invalid Excel file: {str(e)}'}, status=400)

        # Expect columns: Team Name, Category, Season, Group
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        required_cols = ['Team Name', 'Category', 'Season', 'Group']
        if not all(col in header for col in required_cols):
            return Response({'error': f'Missing required columns. Found: {header}'}, status=400)

        results = []
        # cache existing team names for duplicate detection
        existing_names = list(Team.objects.values_list('name', flat=True))
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(header, row))
            team_name = data.get('Team Name')
            category = data.get('Category')
            season_name = data.get('Season')
            group_name = data.get('Group')
            if not all([team_name, category, season_name, group_name]):
                results.append({'team': team_name, 'status': 'skipped', 'reason': 'Missing data'})
                continue

            # detect near-duplicate team names to avoid creating duplicates
            close = difflib.get_close_matches(str(team_name).strip(), existing_names, n=3, cutoff=0.82)
            if close:
                results.append({'team': team_name, 'status': 'duplicate_suspected', 'reason': 'Similar team exists', 'suggestions': close})
                continue

            season = Season.objects.filter(name=season_name, category=category).first()
            if not season:
                results.append({'team': team_name, 'status': 'skipped', 'reason': 'Season not found'})
                continue

            group, _ = Group.objects.get_or_create(name=group_name, season=season)
            team, _ = Team.objects.get_or_create(name=team_name, defaults={'category': category})
            TeamGroup.objects.update_or_create(team=team, group=group)
            results.append({'team': team_name, 'status': 'imported'})

        return Response({'results': results})

@api_view(['GET'])
def standings_view(request, season_id=None):
    """Return standings for a given season id, or if not provided pick the latest season.

    Query params supported:
    - season_id: explicit season id (path param takes precedence)
    - category: if season_id not provided, pick latest season for this category
    """
    # If called with path param, DRF will pass season_id; otherwise check query params
    sid = season_id or request.query_params.get('season_id')
    if sid:
        try:
            sid = int(sid)
        except (TypeError, ValueError):
            return Response({'error': 'Invalid season_id'}, status=400)
        data = compute_standings(sid)
        return Response(data)

    # No season id provided — try to pick latest by category or overall
    category = request.query_params.get('category')
    if category:
        season = Season.objects.filter(category=category).order_by('-start_date').first()
    else:
        season = Season.objects.order_by('-start_date').first()

    if not season:
        return Response({'error': 'No season found'}, status=404)

    data = compute_standings(season.id)
    return Response(data)

class TeamViewSet(viewsets.ModelViewSet):
    queryset = Team.objects.filter(archived=False)
    serializer_class = TeamSerializer

class SeasonViewSet(viewsets.ModelViewSet):
    queryset = Season.objects.all().order_by('-start_date')
    serializer_class = SeasonSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        category = self.request.query_params.get('category')
        if category:
            qs = qs.filter(category=category)
        return qs

class MatchViewSet(viewsets.ModelViewSet):
    # Include archived teams (used for bracket placeholders) but filter them when not showing brackets
    queryset = Match.objects.select_related('home_team', 'away_team')
    # Use different serializer for read vs write
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return MatchCreateSerializer
        return MatchSerializer

    def get_queryset(self):
        # Start from base queryset (includes related teams)
        qs = self.queryset
        # Allow optional filtering by season id, season_name, or category via query params
        season = self.request.query_params.get('season')
        season_name = self.request.query_params.get('season_name')
        category = self.request.query_params.get('category')

        if season:
            try:
                sid = int(season)
                qs = qs.filter(season_id=sid)
            except (TypeError, ValueError):
                pass
        elif season_name:
            qs = qs.filter(season__name__iexact=season_name)
        elif category:
            qs = qs.filter(season__category=category)

        # Order matches by matchday then match_date so bracket rendering is stable
        return qs.order_by('matchday', 'match_date')

    # extra: endpoint to mark a match result
    @action(detail=True, methods=['post'], permission_classes=[IsResultsEditor])
    def set_result(self, request, pk=None):
        match = self.get_object()
        home_score = request.data.get('home_score')
        away_score = request.data.get('away_score')
        # optional period param from frontend
        period = request.data.get('period') or request.data.get('current_period')
        if home_score is None or away_score is None:
            return Response({'error': 'Provide home_score and away_score'}, status=400)
        try:
            match.home_score = int(home_score)
            match.away_score = int(away_score)
        except (TypeError, ValueError):
            return Response({'error': 'Scores must be integers'}, status=400)
        # If actual_start is not set, record it when results are first saved
        if not match.actual_start:
            try:
                match.actual_start = timezone.now()
            except Exception:
                pass
        # persist a provided period string (e.g. '1st_half', 'halftime', '2nd_half')
        if period:
            try:
                match.current_period = period
            except Exception:
                pass
        # When saving interim scores we do NOT want to auto-mark the match as
        # completed. Use a temporary instance flag that Match.save() honors
        # to suppress automatic is_played calculation.
        match._suppress_auto_played = True
        match.save()
        return Response(self.get_serializer(match).data)

    @action(detail=True, methods=['post'], permission_classes=[IsResultsEditor])
    def set_period(self, request, pk=None):
        """Explicit endpoint to update the `current_period` for a match.

        Accepts JSON: { "period": "1st_half" } or { "current_period": "halftime" }.
        Validates against model choices and returns the updated match.
        """
        match = self.get_object()
        period = request.data.get('period') or request.data.get('current_period')
        if not period:
            return Response({'error': 'Provide a period value'}, status=status.HTTP_400_BAD_REQUEST)
        # validate against model choices
        allowed = [k for k, _ in Match.CURRENT_PERIOD_CHOICES]
        if period not in allowed:
            return Response({'error': 'Invalid period value'}, status=status.HTTP_400_BAD_REQUEST)
        match.current_period = period
        match.save()
        return Response(self.get_serializer(match).data)

    @action(detail=True, methods=['post'], permission_classes=[IsResultsEditor])
    def mark_finished(self, request, pk=None):
        """Mark a match as finished manually. Accepts optional `extra_time_minutes` (int).
        If not provided, backend will pick a random 1-5 minutes. Records who marked it and timestamp.
        """
        match = self.get_object()
        extra = request.data.get('extra_time_minutes')
        try:
            if extra is not None and extra != '':
                extra_val = int(extra)
            else:
                extra_val = random.randint(1, 5)
        except (TypeError, ValueError):
            return Response({'error': 'extra_time_minutes must be an integer'}, status=400)

        match.manual_finished_at = timezone.now()
        match.extra_time_minutes = extra_val
        user = request.user
        match.manual_finished_by = user.username if (user and getattr(user, 'is_authenticated', False)) else ''
        match.is_played = True
        match.save()
        return Response(self.get_serializer(match).data)

class NewsViewSet(viewsets.ModelViewSet):
    queryset = News.objects.all().order_by('-published_at')
    serializer_class = NewsSerializer
    permission_classes = [IsNewsUploaderOrReadOnly]


# Create your views here.


@api_view(['GET'])
def groups_with_teams(request):
    season_id = request.query_params.get('season')
    if not season_id:
        return Response({'error': 'season parameter required'}, status=status.HTTP_400_BAD_REQUEST)
    groups = Group.objects.filter(season_id=season_id).order_by('name')
    data = []
    for g in groups:
        teams = [
            {'id': tg.team.id, 'name': tg.team.name}
            for tg in TeamGroup.objects.filter(group=g, team__archived=False).select_related('team')
        ]
        data.append({'id': g.id, 'name': g.name, 'teams': teams})
    return Response(data)


@api_view(['GET'])
def grouped_standings(request):
    season_id = request.query_params.get('season')
    if not season_id:
        return Response({'error': 'season parameter required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        sid = int(season_id)
    except (TypeError, ValueError):
        return Response({'error': 'invalid season id'}, status=status.HTTP_400_BAD_REQUEST)

    # compute overall standings
    season_standings = compute_standings(sid)

    # get groups for season
    groups = Group.objects.filter(season_id=sid).order_by('name')
    result = []
    for g in groups:
        team_ids = [tg.team.id for tg in TeamGroup.objects.filter(group=g).select_related('team')]
        # filter standings for teams in this group
        group_rows = [row for row in season_standings if row['team_id'] in team_ids]
        # sort using same comparator as compute_standings (already sorted overall, but ensure group order)
        group_rows.sort(key=lambda x: (-x['points'], -x['goal_diff'], -x['goals_for'], x['team_name']))
        result.append({'group': {'id': g.id, 'name': g.name}, 'standings': group_rows})

    return Response(result)


@api_view(['GET'])
def me(request):
    """Return basic info about the authenticated user including groups."""
    user = request.user
    if not user or not user.is_authenticated:
        return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
    groups = list(user.groups.values_list('name', flat=True))
    return Response({
        'username': user.username,
        'is_superuser': user.is_superuser,
        'groups': groups,
    })


@api_view(['POST'])
def move_team(request):
    team_id = request.data.get('team_id')
    to_group_id = request.data.get('to_group_id')
    season_id = request.data.get('season_id')
    if not all([team_id, to_group_id, season_id]):
        return Response({'error': 'team_id, to_group_id, and season_id required'}, status=status.HTTP_400_BAD_REQUEST)
    team = get_object_or_404(Team, id=team_id)
    to_group = get_object_or_404(Group, id=to_group_id, season_id=season_id)
    # Ensure team removed from any existing group for this season
    TeamGroup.objects.filter(team=team, season_id=season_id).delete()
    tg = TeamGroup.objects.create(team=team, group=to_group, season_id=season_id)
    return Response({'success': True, 'team': team.name, 'group': to_group.name})


@api_view(['GET'])
def export_played_matches(request):
    """Export played matches for a season as CSV attachment.

    Query params:
      - season: numeric season id (preferred)
      - season_name: season name fallback

    Only staff users may download exports.
    """
    user = request.user
    if not (user and getattr(user, 'is_authenticated', False) and getattr(user, 'is_staff', False)):
        return Response({'error': 'staff access required'}, status=403)

    season_param = request.query_params.get('season')
    season_name = request.query_params.get('season_name')
    season_obj = None
    if season_param:
        try:
            season_id = int(season_param)
            from .models import Season
            season_obj = Season.objects.filter(id=season_id).first()
        except Exception:
            season_obj = None
    elif season_name:
        from .models import Season
        season_obj = Season.objects.filter(name__iexact=season_name).first()

    if not season_obj:
        return Response({'error': 'season not found; provide season or season_name'}, status=400)

    from django.http import HttpResponse
    import csv

    matches = Match.objects.filter(season=season_obj, is_played=True).select_related('home_team', 'away_team', 'season', 'awarded_to').order_by('matchday', 'match_date')

    # prepare CSV
    filename = f"matches_played_season_{season_obj.id}.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    header = ['id', 'season_id', 'season_name', 'matchday', 'match_date', 'venue', 'home_team_id', 'home_team_name', 'away_team_id', 'away_team_name', 'home_score', 'away_score', 'penalty_home', 'penalty_away', 'is_played', 'awarded', 'awarded_reason', 'awarded_to_id', 'awarded_to_name', 'manual_finished_at', 'actual_start', 'current_period', 'match_duration_minutes']
    writer.writerow(header)

    for m in matches:
        row = [
            m.id,
            m.season_id,
            getattr(m.season, 'name', ''),
            m.matchday,
            (m.match_date.isoformat() if m.match_date else ''),
            m.venue,
            m.home_team_id,
            (m.home_team.name if hasattr(m.home_team, 'name') else m.home_team),
            m.away_team_id,
            (m.away_team.name if hasattr(m.away_team, 'name') else m.away_team),
            m.home_score,
            m.away_score,
            m.penalty_home,
            m.penalty_away,
            m.is_played,
            getattr(m, 'awarded', False),
            getattr(m, 'awarded_reason', ''),
            (m.awarded_to_id if hasattr(m, 'awarded_to_id') else (m.awarded_to.id if m.awarded_to else None)),
            (m.awarded_to.name if getattr(m, 'awarded_to', None) and getattr(m.awarded_to, 'name', None) else ''),
            (m.manual_finished_at.isoformat() if m.manual_finished_at else ''),
            (m.actual_start.isoformat() if m.actual_start else ''),
            getattr(m, 'current_period', ''),
            getattr(m, 'match_duration_minutes', ''),
        ]
        writer.writerow(row)

    return response
